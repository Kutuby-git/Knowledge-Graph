"""
Quranic Arabic Words - Excel to Supabase Ingestion Pipeline

Parses the master Excel file with 16 sheets (15 units + 1 Advanced),
handles 3 different sheet formats, fixes Excel Surah:Ayah corruption,
deduplicates by bare Arabic text, and uploads to Supabase.

Expected counts: 768 unique words, 15 units, ~130 sub_themes, 988 word_theme records
"""

import datetime
import os
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(Path(__file__).parent.parent / ".env.local")

SUPABASE_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]

EXCEL_PATH = Path(__file__).parent.parent / "Quranic Arabic Words  - MASTER LIST (1).xlsx"

# Sheet name -> unit code mapping
UNIT_SHEETS = {
    "Unit A Animals in the Quran": "A",
    "Unit B Fruits, Plants, Creation": "B",
    "Unit C People and Family": "C",
    "Unit D Places and Nature": "D",
    "Unit E Names of Allah": "E",
    "Unit F Salah and Worship": "F",
    "Unit G Time, Numbers, Days": "G",
    "Unit H Character and Ethics": "H",
    "Unit I Body and Senses": "I",
}

# These sheet names may be truncated in Excel
UNIT_SHEETS_FUZZY = {
    "J": "Opposites",
    "K": "Prophets",
    "L": "Virtue",
    "M": "Spiritual",
    "N": "Emotions",
    "O": "Colors",
}

# Sheet format classification
DUAL_15_COL = {"G", "H"}  # 15-column dual layout
DUAL_18_COL = {"E"}        # 18-column dual layout (Names of Allah)
STANDARD = {"A", "B", "C", "D", "F", "I", "J", "K", "L", "M", "N", "O"}


def strip_diacritics(text: str) -> str:
    """Remove Arabic diacritics to get bare consonantal form for deduplication."""
    if not text:
        return ""
    # Arabic diacritics Unicode range
    diacritics = set([
        '\u0610', '\u0611', '\u0612', '\u0613', '\u0614', '\u0615',
        '\u0616', '\u0617', '\u0618', '\u0619', '\u061A',
        '\u064B', '\u064C', '\u064D', '\u064E', '\u064F', '\u0650',
        '\u0651', '\u0652', '\u0653', '\u0654', '\u0655', '\u0656',
        '\u0657', '\u0658', '\u0659', '\u065A', '\u065B', '\u065C',
        '\u065D', '\u065E', '\u065F', '\u0670',
    ])
    return "".join(c for c in text if c not in diacritics).strip()


def convert_surah_ayah(value) -> str | None:
    """Convert Excel-corrupted Surah:Ayah values back to 'X:Y' strings."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        return value
    if isinstance(value, datetime.time):
        return f"{value.hour}:{value.minute}"
    if isinstance(value, datetime.timedelta):
        total_seconds = int(value.total_seconds())
        surah = total_seconds // 3600
        ayah = (total_seconds % 3600) // 60
        return f"{surah}:{ayah}"
    if isinstance(value, (int, float)):
        total_hours = float(value) * 24
        surah = int(total_hours)
        ayah = int(round((total_hours - surah) * 60))
        return f"{surah}:{ayah}"
    return str(value)


def is_arabic(text: str) -> bool:
    """Check if text contains Arabic characters."""
    if not text:
        return False
    for char in str(text):
        if '\u0600' <= char <= '\u06FF' or '\u0750' <= char <= '\u077F':
            return True
    return False


def clean_str(val) -> str | None:
    """Convert cell value to clean string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_standard_sheet(ws, unit_code: str) -> list[dict]:
    """Parse standard 6-column sheet format (Units A,B,C,D,F,I,J,K,L,M,N,O)."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    entries = []

    for row_idx, row in enumerate(rows, start=2):
        if not row or len(row) < 6:
            continue

        # Unit K: detect shifted columns (rows 83-86)
        if unit_code == "K" and len(row) >= 7:
            col3_val = clean_str(row[2])
            if col3_val and not is_arabic(col3_val) and is_arabic(str(row[3] or "")):
                # Shifted: col 3 is sub-label, actual data starts at col 3 (0-indexed)
                arabic = clean_str(row[3])
                if not arabic or not is_arabic(arabic):
                    continue
                entries.append({
                    "theme_order": row[0],
                    "theme_label": clean_str(row[1]),
                    "arabic": arabic,
                    "transliteration": clean_str(row[4]),
                    "kids_glossary": clean_str(row[5]),
                    "surah_ayah": convert_surah_ayah(row[6]),
                    "part_2": False,
                    "source_sheet": f"Unit {unit_code}",
                    "row_number": row_idx,
                })
                continue

        arabic = clean_str(row[2])
        if not arabic or not is_arabic(arabic):
            continue

        part_2 = False
        if len(row) >= 7 and row[6]:
            part_2 = str(row[6]).strip().lower() in ("yes", "true", "1")

        entries.append({
            "theme_order": row[0],
            "theme_label": clean_str(row[1]),
            "arabic": arabic,
            "transliteration": clean_str(row[3]),
            "kids_glossary": clean_str(row[4]),
            "surah_ayah": convert_surah_ayah(row[5]),
            "part_2": part_2,
            "source_sheet": f"Unit {unit_code}",
            "row_number": row_idx,
        })

    return entries


def parse_dual_15_sheet(ws, unit_code: str) -> list[dict]:
    """Parse 15-column dual-layout sheet (Units G, H). Right side is authoritative."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    entries = []

    for row_idx, row in enumerate(rows, start=2):
        if not row or len(row) < 15:
            continue

        # Right side columns: 8-14 (0-indexed)
        arabic = clean_str(row[10])  # col 11 = Arabic (right side)
        if not arabic or not is_arabic(arabic):
            continue

        part_2 = False
        if row[14]:
            part_2 = str(row[14]).strip().lower() in ("yes", "true", "1")

        entries.append({
            "theme_order": row[8],    # col 9
            "theme_label": clean_str(row[9]),  # col 10
            "arabic": arabic,
            "transliteration": clean_str(row[11]),  # col 12
            "kids_glossary": clean_str(row[12]),     # col 13
            "surah_ayah": convert_surah_ayah(row[13]),  # col 14
            "part_2": part_2,
            "source_sheet": f"Unit {unit_code}",
            "row_number": row_idx,
        })

    return entries


def parse_dual_18_sheet(ws, unit_code: str) -> list[dict]:
    """Parse 18-column dual-layout sheet (Unit E - Names of Allah). Right side authoritative."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    entries = []

    for row_idx, row in enumerate(rows, start=2):
        if not row or len(row) < 18:
            continue

        # Right side columns: 11-17 (0-indexed)
        arabic = clean_str(row[13])  # col 14 = Arabic
        if not arabic or not is_arabic(arabic):
            continue

        part_2 = False
        if row[17]:
            part_2 = str(row[17]).strip().lower() in ("yes", "true", "1")

        entries.append({
            "theme_order": row[11],   # col 12
            "theme_label": clean_str(row[12]),  # col 13
            "arabic": arabic,
            "transliteration": clean_str(row[14]),  # col 15
            "kids_glossary": clean_str(row[15]),     # col 16
            "surah_ayah": convert_surah_ayah(row[16]),  # col 17
            "part_2": part_2,
            "source_sheet": f"Unit {unit_code}",
            "row_number": row_idx,
        })

    return entries


def parse_advanced_sheet(ws) -> list[dict]:
    """Parse Advanced sheet with unit-coded ThemeOrders like 'A8', 'B6'."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    entries = []

    for row_idx, row in enumerate(rows, start=2):
        if not row or len(row) < 6:
            continue

        theme_order_raw = clean_str(row[0])
        if not theme_order_raw:
            continue

        arabic = clean_str(row[2])
        if not arabic or not is_arabic(arabic):
            continue

        # Parse unit-coded theme order: "A8" -> unit="A", order=8
        match = re.match(r"^([A-O])(\d+)$", theme_order_raw)
        if not match:
            continue

        unit_code = match.group(1)
        theme_num = float(match.group(2))

        entries.append({
            "unit_code": unit_code,
            "theme_order": theme_num,
            "theme_label": clean_str(row[1]),
            "arabic": arabic,
            "transliteration": clean_str(row[3]),
            "kids_glossary": clean_str(row[4]),
            "surah_ayah": convert_surah_ayah(row[5]),
            "part_2": False,
            "is_advanced": True,
            "source_sheet": "Advanced",
            "row_number": row_idx,
        })

    return entries


def find_sheet(wb, unit_code: str):
    """Find worksheet by unit code, handling truncated sheet names."""
    # Try exact match first
    for name, code in UNIT_SHEETS.items():
        if code == unit_code and name in wb.sheetnames:
            return wb[name]

    # Try fuzzy match for J-O
    keyword = UNIT_SHEETS_FUZZY.get(unit_code)
    if keyword:
        for sheet_name in wb.sheetnames:
            if f"Unit {unit_code}" in sheet_name or keyword in sheet_name:
                return wb[sheet_name]

    return None


def main():
    print("=== Quranic Arabic Words Ingestion Pipeline ===\n")

    # Connect to Supabase
    print("Connecting to Supabase...")
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Load Excel
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Sheets found: {wb.sheetnames}\n")

    # Fetch unit records from DB (already seeded by migration)
    units_resp = sb.table("units").select("*").execute()
    units_by_code = {u["code"]: u for u in units_resp.data}
    print(f"Units in DB: {len(units_by_code)}")

    # ========================================
    # STEP 1: Parse all sheets
    # ========================================
    all_entries = []  # List of dicts with unit_code added

    for unit_code in "ABCDEFGHIJKLMNO":
        ws = find_sheet(wb, unit_code)
        if not ws:
            print(f"  WARNING: Sheet for Unit {unit_code} not found!")
            continue

        if unit_code in DUAL_18_COL:
            entries = parse_dual_18_sheet(ws, unit_code)
        elif unit_code in DUAL_15_COL:
            entries = parse_dual_15_sheet(ws, unit_code)
        else:
            entries = parse_standard_sheet(ws, unit_code)

        for e in entries:
            e.setdefault("unit_code", unit_code)
            e.setdefault("is_advanced", False)

        all_entries.extend(entries)
        print(f"  Unit {unit_code}: {len(entries)} entries parsed")

    # Parse Advanced sheet
    adv_sheet = None
    for name in wb.sheetnames:
        if "advanced" in name.lower():
            adv_sheet = wb[name]
            break

    adv_entries = []
    if adv_sheet:
        adv_entries = parse_advanced_sheet(adv_sheet)
        all_entries.extend(adv_entries)
        print(f"  Advanced: {len(adv_entries)} entries parsed")
    else:
        print("  WARNING: Advanced sheet not found!")

    print(f"\nTotal entries (with duplicates): {len(all_entries)}")

    # ========================================
    # STEP 2: Collect sub-themes and insert
    # ========================================
    print("\nInserting sub-themes...")
    sub_themes_seen = {}  # (unit_code, theme_order) -> label

    for entry in all_entries:
        uc = entry["unit_code"]
        to = entry["theme_order"]
        label = entry["theme_label"]
        if to is None or label is None:
            continue
        to_float = float(to)
        key = (uc, to_float)
        if key not in sub_themes_seen:
            sub_themes_seen[key] = label

    # Insert sub-themes
    sub_theme_records = []
    for (uc, to), label in sorted(sub_themes_seen.items()):
        unit = units_by_code.get(uc)
        if not unit:
            print(f"  WARNING: No unit found for code {uc}")
            continue
        sub_theme_records.append({
            "unit_id": unit["id"],
            "theme_order": to,
            "label": label,
            "display_order": int(to),
        })

    if sub_theme_records:
        # Clear existing and insert fresh
        sb.table("sub_themes").delete().neq("id", 0).execute()
        # Insert in batches of 50
        for i in range(0, len(sub_theme_records), 50):
            batch = sub_theme_records[i:i+50]
            sb.table("sub_themes").insert(batch).execute()

    # Refetch sub_themes with IDs
    st_resp = sb.table("sub_themes").select("*, units(code)").execute()
    st_lookup = {}
    for st in st_resp.data:
        unit_code = st["units"]["code"]
        key = (unit_code, float(st["theme_order"]))
        st_lookup[key] = st["id"]

    print(f"  Sub-themes inserted: {len(sub_theme_records)}")

    # ========================================
    # STEP 3: Deduplicate words and insert
    # ========================================
    print("\nDeduplicating and inserting words...")

    # Group entries by bare Arabic text
    words_map = {}  # arabic_bare -> {word_data, theme_associations: [...]}

    for entry in all_entries:
        arabic = entry["arabic"]
        if not arabic:
            continue
        bare = strip_diacritics(arabic)
        if not bare:
            continue

        if bare not in words_map:
            words_map[bare] = {
                "arabic": arabic,
                "arabic_bare": bare,
                "transliteration": entry["transliteration"],
                "kids_glossary": entry["kids_glossary"],
                "surah_ayah": entry["surah_ayah"],
                "is_advanced": entry.get("is_advanced", False),
                "part_2": entry.get("part_2", False),
                "associations": [],
            }
        else:
            # Merge: prefer non-null values, keep advanced flag if any source is advanced
            existing = words_map[bare]
            if not existing["transliteration"] and entry["transliteration"]:
                existing["transliteration"] = entry["transliteration"]
            if not existing["kids_glossary"] and entry["kids_glossary"]:
                existing["kids_glossary"] = entry["kids_glossary"]
            if not existing["surah_ayah"] and entry["surah_ayah"]:
                existing["surah_ayah"] = entry["surah_ayah"]
            if entry.get("is_advanced"):
                existing["is_advanced"] = True
            if entry.get("part_2"):
                existing["part_2"] = True

        # Track theme association
        uc = entry["unit_code"]
        to = entry["theme_order"]
        if to is not None:
            st_key = (uc, float(to))
            st_id = st_lookup.get(st_key)
            if st_id:
                words_map[bare]["associations"].append({
                    "sub_theme_id": st_id,
                    "source_sheet": entry.get("source_sheet"),
                    "row_number": entry.get("row_number"),
                })

    print(f"  Unique words: {len(words_map)}")

    # Clear existing words
    sb.table("word_themes").delete().neq("id", 0).execute()
    sb.table("words").delete().neq("id", 0).execute()

    # Insert words in batches
    word_records = []
    for bare, data in words_map.items():
        word_records.append({
            "arabic": data["arabic"],
            "arabic_bare": data["arabic_bare"],
            "transliteration": data["transliteration"],
            "kids_glossary": data["kids_glossary"],
            "surah_ayah": data["surah_ayah"],
            "is_advanced": data["is_advanced"],
            "part_2": data["part_2"],
        })

    inserted_words = []
    for i in range(0, len(word_records), 50):
        batch = word_records[i:i+50]
        resp = sb.table("words").insert(batch).execute()
        inserted_words.extend(resp.data)

    # Build lookup: arabic_bare -> word_id
    word_id_lookup = {w["arabic_bare"]: w["id"] for w in inserted_words}
    print(f"  Words inserted: {len(inserted_words)}")

    # ========================================
    # STEP 4: Insert word_themes associations
    # ========================================
    print("\nInserting word-theme associations...")
    wt_records = []
    seen_wt = set()

    for bare, data in words_map.items():
        word_id = word_id_lookup.get(bare)
        if not word_id:
            continue
        for assoc in data["associations"]:
            key = (word_id, assoc["sub_theme_id"])
            if key in seen_wt:
                continue
            seen_wt.add(key)
            wt_records.append({
                "word_id": word_id,
                "sub_theme_id": assoc["sub_theme_id"],
                "source_sheet": assoc["source_sheet"],
                "row_number": assoc["row_number"],
            })

    for i in range(0, len(wt_records), 50):
        batch = wt_records[i:i+50]
        sb.table("word_themes").insert(batch).execute()

    print(f"  Word-theme associations: {len(wt_records)}")

    # ========================================
    # STEP 5: Seed initial relationships
    # ========================================
    print("\nSeeding initial relationships...")
    rel_count = 0

    # Find opposite pairs from Unit J (Opposites and Big Concepts)
    # Words in Unit J are typically organized as pairs within the same sub-theme
    unit_j = units_by_code.get("J")
    if unit_j:
        j_themes = sb.table("sub_themes").select("id").eq("unit_id", unit_j["id"]).execute()
        for st in j_themes.data:
            wt_resp = sb.table("word_themes").select("word_id").eq("sub_theme_id", st["id"]).execute()
            word_ids = [wt["word_id"] for wt in wt_resp.data]
            # Pair consecutive words as opposites
            for k in range(0, len(word_ids) - 1, 2):
                try:
                    sb.table("word_relationships").insert({
                        "word_id_1": word_ids[k],
                        "word_id_2": word_ids[k + 1],
                        "relationship_type": "opposite",
                        "weight": 0.8,
                    }).execute()
                    rel_count += 1
                except Exception:
                    pass  # Skip if duplicate

    # Find virtue pairs from Unit L
    unit_l = units_by_code.get("L")
    if unit_l:
        l_themes = sb.table("sub_themes").select("id").eq("unit_id", unit_l["id"]).execute()
        for st in l_themes.data:
            wt_resp = sb.table("word_themes").select("word_id").eq("sub_theme_id", st["id"]).execute()
            word_ids = [wt["word_id"] for wt in wt_resp.data]
            for k in range(0, len(word_ids) - 1, 2):
                try:
                    sb.table("word_relationships").insert({
                        "word_id_1": word_ids[k],
                        "word_id_2": word_ids[k + 1],
                        "relationship_type": "virtue_pair",
                        "weight": 0.8,
                    }).execute()
                    rel_count += 1
                except Exception:
                    pass

    print(f"  Relationships seeded: {rel_count}")

    # ========================================
    # SUMMARY
    # ========================================
    print("\n=== INGESTION COMPLETE ===")
    print(f"  Units: {len(units_by_code)}")
    print(f"  Sub-themes: {len(sub_theme_records)}")
    print(f"  Unique words: {len(inserted_words)}")
    print(f"  Word-theme associations: {len(wt_records)}")
    print(f"  Relationships: {rel_count}")

    # Validate
    words_count = sb.table("words").select("id", count="exact").execute()
    print(f"\n  DB word count: {words_count.count}")

    # Check Surah:Ayah format
    invalid = sb.table("words").select("id, surah_ayah").not_.is_("surah_ayah", "null").execute()
    bad_refs = [w for w in invalid.data if w["surah_ayah"] and not re.match(r"^\d+:\d+$", w["surah_ayah"])]
    if bad_refs:
        print(f"  WARNING: {len(bad_refs)} words with invalid Surah:Ayah format")
        for b in bad_refs[:5]:
            print(f"    id={b['id']}: '{b['surah_ayah']}'")
    else:
        print("  All Surah:Ayah values valid ✓")


if __name__ == "__main__":
    main()
